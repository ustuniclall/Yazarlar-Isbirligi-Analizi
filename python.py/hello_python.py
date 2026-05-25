#Bütün exceli çalıştırıyo ama kenar ağırlıklarında sıkınıt  var
import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl import load_workbook
import networkx as nx
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk
import openpyxl
import scipy as sp
from collections import defaultdict
import pandas as pd
from queue import PriorityQueue

from anytree import Node, RenderTree
from anytree.exporter import DotExporter
from tkinter import simpledialog
import heapq


class GraphVisualization:
    def __init__(self, graph, sol_panel):
        self.graph = graph
        self.sol_panel = sol_panel

    def yazarlar_arasi_alt_graf(self):
        # Kullanıcıdan yazar ID'leri alınır
        yazar_a = simpledialog.askstring("Yazar A ID", "Lütfen A yazarının ID'sini giriniz:")
        yazar_b = simpledialog.askstring("Yazar B ID", "Lütfen B yazarının ID'sini giriniz:")

        if not yazar_a or not yazar_b:
            return

        if yazar_a not in self.graph.nodes or yazar_b not in self.graph.nodes:
            messagebox.showerror("Hata", f"Hata: '{yazar_a}' veya '{yazar_b}' grafikte mevcut değil!")
            return

        try:
            # A ve B arasındaki en kısa yol ve mesafeyi hesapla
            yol = nx.dijkstra_path(self.graph, source=yazar_a, target=yazar_b)
            mesafe = nx.dijkstra_path_length(self.graph, source=yazar_a, target=yazar_b)
        except nx.NetworkXNoPath:
            messagebox.showerror("Hata", f"Hata: '{yazar_a}' ile '{yazar_b}' arasında bir bağlantı yok!")
            return

        # Kuyruk içeriğini adım adım göstermek için PriorityQueue simülasyonu
        pq = PriorityQueue()
        pq.put((0, [yazar_a]))

        # Kuyruk adımlarını listeye atıyoruz
        adimlar = []
        while not pq.empty():
            maliyet, yol_sozde = pq.get()
            adimlar.append(f"Maliyet: {maliyet}, Yol: {yol_sozde}")

            if yol_sozde[-1] == yazar_b:
                break

            for komsu, veri in self.graph[yol_sozde[-1]].items():
                if komsu not in yol_sozde:
                    yeni_maliyet = maliyet + veri['weight']
                    yeni_yol = yol_sozde + [komsu]
                    pq.put((yeni_maliyet, yeni_yol))

        # Kuyruk adımlarını messagebox ile gösterme
        adimlar_mesaji = "\n".join(adimlar)
        messagebox.showinfo("Kuyruk Adımları", adimlar_mesaji)

        # Alt graf oluşturma
        alt_graf = nx.compose(self.graph.subgraph(nx.single_source_shortest_path(self.graph, source=yazar_a).keys()),
                              self.graph.subgraph(nx.single_source_shortest_path(self.graph, source=yazar_b).keys()))

        # En kısa yollar ve mesafeler
        yollar_a = nx.single_source_dijkstra_path(alt_graf, source=yazar_a)
        yollar_b = nx.single_source_dijkstra_path(alt_graf, source=yazar_b)

        yollar = {**yollar_a, **yollar_b}
        mesafeler = nx.single_source_dijkstra_path_length(alt_graf, source=yazar_a)

        # Renkli görselleştirme
        renkler = []
        for node in alt_graf.nodes:
            if node in yollar_a and node in yollar_b:
                renkler.append('orchid')
            elif node == yazar_a or node == yazar_b:
                renkler.append('purple')
            else:
                renkler.append('red')

        # Görselleştirme
        pos = nx.spring_layout(alt_graf)
        nx.draw(
            alt_graf, pos, with_labels=True, node_color=renkler,
            node_size=500, font_size=10, font_color='black'
        )
        plt.title(f"'{yazar_a}' ve '{yazar_b}' Yazarlarının Bağlantılarıyla Olan En Kısa Yollar")
        plt.interactive(True) 
        plt.show()


# Excel dosyasından graf oluşturma fonksiyonu
def excelden_graf_olusturma(file_name):
    wb = load_workbook(filename=file_name)
    sheet = wb.active
    
    coauthors_sutun_index = 5  # "coauthors" E sütununda olduğu için indeks 5
    graph = nx.Graph()
    article_counts = {}  # Yazarların yazdığı makale sayıları

    for satir in sheet.iter_rows(min_row=2, values_only=True):  # Excel'deki tüm satırlar okunuyor
        coauthors = satir[coauthors_sutun_index - 1]  # E sütunu için indeks 5 -> 5-1 = 4
        if coauthors:
            authors = [author.replace("[", "").replace("]", "").replace("'", "").replace(",", "").strip() for author in coauthors.split(",")]

            for author in authors:
                if author not in article_counts:
                    article_counts[author] = 0
                article_counts[author] += 1

            if len(authors) > 1:
                for i, author1 in enumerate(authors):
                    for author2 in authors[i + 1:]:
                        if graph.has_edge(author1, author2):
                            graph[author1][author2]['weight'] += 1
                        else:
                            graph.add_edge(author1, author2, weight=1)

    if len(graph.nodes) == 0:
        raise ValueError("Excel dosyasından graf oluşturulamadı. Verileri kontrol edin.")
    
    return graph, article_counts
           

def update_output(ax_output, text):
    ax_output.clear()
    ax_output.text(0.05, 0.5, text, fontsize=10, ha="left", va="center", wrap=True,
                   bbox=dict(facecolor="lightgrey", edgecolor="black"))
    plt.draw()


# Dijkstra algoritması ile en kısa yolu bulma fonksiyonu
def dijkstra(graph, baslangic, bitis):
    kuyruk = [(0, baslangic)]  # (mesafe, yazar)
    mesafeler = {baslangic: 0}
    onceki_dugum = {baslangic: None}
    adimlar = []  # Adım adım kuyruğu göstermek için

    while kuyruk:
        simdiki_mesafe, simdiki_dugum = heapq.heappop(kuyruk)
        adimlar.append((list(kuyruk), simdiki_dugum))  # Kuyruğun içeriği ve şu anki yazar

        if simdiki_dugum == bitis:
            # Yol bulunduysa, geri gidip yolu oluşturuyoruz
            yol = []
            while onceki_dugum[simdiki_dugum] is not None:
                yol.insert(0, simdiki_dugum)
                simdiki_dugum = onceki_dugum[simdiki_dugum]
            yol.insert(0, baslangic)  # Başlangıç noktasını da ekleyelim
            return yol, adimlar

        for komsu, weight in graph.get(simdiki_dugum, {}).items():
            mesafe = simdiki_mesafe + weight['weight']
            if komsu not in mesafeler or mesafe < mesafeler[komsu]:
                mesafeler[komsu] = mesafe
                onceki_dugum[komsu] = simdiki_dugum
                heapq.heappush(kuyruk, (mesafe, komsu))

    return None, adimlar  # Eğer yol bulunamazsa

# Düğüm boyutlarını ve renklerini ayarlama fonksiyonu
def dugum_duzeni(graph, article_counts):
    avg_articles = sum(article_counts.values()) / len(article_counts)

    dugum_boyutu = []
    dugum_renkleri = []

    for dugum in graph.nodes:
        article_count = article_counts.get(dugum, 0)

        if article_count > avg_articles * 1.2:
            dugum_boyutu.append(800)
            dugum_renkleri.append('purple')
        elif article_count < avg_articles * 0.8:
            dugum_boyutu.append(200)
            dugum_renkleri.append('orchid')
        else:
            dugum_boyutu.append(400)
            dugum_renkleri.append('lightblue')

    return dugum_boyutu, dugum_renkleri

# Kenarları sabit renk ile ayarlama fonksiyonu
def kenar_duzeni(graph):
    kenar_renkleri = ['gray'] * len(graph.edges())
    kenar_genislik = [1] * len(graph.edges())
    return kenar_renkleri, kenar_genislik

# Grafı görselleştirme fonksiyonu
def graf_gorsellestirme(graph, article_counts, panel):
    bicim = nx.spring_layout(graph)
    dugum_boyutu, dugum_renkleri = dugum_duzeni(graph, article_counts)
    kenar_renkleri, kenar_genislik = kenar_duzeni(graph)

    fig, ax = plt.subplots(figsize=(8, 8))
    nx.draw(graph, bicim, with_labels=True, node_size=dugum_boyutu, node_color=dugum_renkleri, edge_color=kenar_renkleri, width=kenar_genislik, ax=ax)

    dugum_etiketleri = {dugum: article_counts[dugum] for dugum in graph.nodes}
    nx.draw_networkx_labels(graph, bicim, labels=dugum_etiketleri, font_size=14, font_color='black', font_weight='bold', verticalalignment='bottom', horizontalalignment='center')

    author_articles = {}
    for satir in sheet.iter_rows(min_row=2, values_only=True):
        author_name, paper_title = satir[3], satir[5]
        if author_name not in author_articles:
          author_articles[author_name] = []
        author_articles[author_name].append(paper_title)

    def tiklama(event):
        if event.button == 1:
            x, y = event.xdata, event.ydata
            tiklanan_dugum = None  # Tıklanan düğümü saklamak için bir değişken

            for dugum, (x_node, y_node) in bicim.items():
                if abs(x - x_node) < 0.1 and abs(y - y_node) < 0.1:
                    tiklanan_dugum = dugum
                    break  # Düğüm bulunduğunda döngüden çık

            if tiklanan_dugum:  # Eğer bir düğüm tıklanmışsa
                yeni_pencere = tk.Toplevel(root)
                metin_kutusu = tk.Text(yeni_pencere)
                metin_kutusu.pack()
                for article in author_articles.get(tiklanan_dugum, []):
                    metin_kutusu.insert(tk.END, article + "\n")

    fig.canvas.mpl_connect('button_press_event', tiklama)
    
    min_agirlik = 3
    edge_labels  = {edge: graph.edges[edge]['weight'] for edge in graph.edges if 'weight' in graph.edges[edge] and graph.edges[edge]['weight'] >= min_agirlik }
    nx.draw_networkx_edge_labels(graph, bicim, edge_labels =edge_labels )

    # Matplotlib grafiklerini Tkinter penceresine yerleştirme
    canvas = FigureCanvasTkAgg(fig, master=panel)  # Tkinter penceresine bağlama
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    canvas.draw()

    # Matplotlib araç çubuğunu ekleme
    toolbar = NavigationToolbar2Tk(canvas, panel)
    toolbar.update()
    toolbar.pack(side=tk.BOTTOM, fill=tk.X)
                      
# Sol panel (Sonuçlar ve detaylar)
class sol_panel(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.text = tk.Text(self, width=40, height=100, wrap=tk.WORD, bg='#ADD8E6')
        self.text.pack(padx=10, pady=10)
        self.pack(side=tk.LEFT)

    def update_text(self, message):
        self.text.insert(tk.END, message + "\n")
        self.text.yview(tk.END)


# Sağ panel (İşlemler)
class sag_panel(tk.Frame):
    def __init__(self, master, left_panel,graph):
        super().__init__(master)
        self.left_panel = left_panel
        self.graph = graph
        self.graph_viz = GraphVisualization(graph,sol_panel) 
        self.pack(side=tk.RIGHT, padx=10)
        

        self.buttons = [
            ("1. İster: En Kısa Yolu Bul", self.graph_viz.yazarlar_arasi_alt_graf),
            ("2. İster: Ağırlığa Göre Kuyruk Oluştur", self.agirliga_gore_kuyruk_olustur),
            ("3. İster: BST Oluştur", self.bst_olustur),
            ("4. İster: Kısa Yolları Hesapla", self.kisa_yollari_hesapla),
            ("5. İster: Toplam İşbirliği Yaptğı Yazar Sayısı", self.toplam_yazar_sayisi_hesapla),
            ("6. İster: En Çok İşbirliği Yapan Yazar", self.en_cok_isbirligi_yapan_yazar),
            ("7. İster: En Uzun Yolu Bul", self.en_uzun_yolu_bul),
        ]

        for text, command in self.buttons:
            button = tk.Button(self, text=text, command=command, width=40, height=6,bg='#ADD8E6')
            button.pack(pady=5, padx=20, anchor="e")

#1.ister

    def yazarlar_arasi_alt_graf(self):
        # Kullanıcıdan A ve B yazarlarının ID'lerini alalım
        A = simpledialog.askstring("Yazar ID'si", "Lütfen Yazar ID'sini giriniz:")
        B = simpledialog.askstring("Yazar ID'si", "Lütfen Yazar ID'sini giriniz:")
        #B = input("B yazarı: ")
        G = graf_gorsellestirme
        # A ve B arasındaki en kısa yolu bulma
        path, distance = dijkstra(G, A, B)

        # En kısa yolu ve mesafeyi yazdırma
        print(f"A ile B arasındaki en kısa yol: {path}")
        print(f"Mesafe: {distance}")

        # En kısa yolu grafik üzerinde gösterme
        pos = nx.spring_layout(G)
        nx.draw(G, pos, with_labels=True, node_size=3000, node_color="lightblue", font_size=15, font_weight="bold")
        #edge_labels = nx.get_edge_attributes(G, 'weight')
        edge_labels = {(u, v): f"{d['weight']}" for u, v, d in graph.edges(data=True)}
        nx.draw_networkx_edge_labels(G, pos=nx.spring_layout(G), edge_labels=edge_labels)

        # En kısa yolu vurgulama
        path_edges = [(path[i], path[i+1]) for i in range(len(path)-1)]
        nx.draw_networkx_edges(G, pos, edgelist=path_edges, edge_color="red", width=2)

        plt.title(f"En Kısa Yol: {A} ile {B}")
        plt.show()

        # Kuyruğu adım adım göstermek için fonksiyon
        def print_queue_steps(graph, start, end):
            queue = [(0, start)]
            distances = {start: 0}
            previous_nodes = {start: None}
            print(f"Başlangıç: {start}")
            print("Adım Adım Kuyruk:")
            while queue:
                current_distance, current_node = heapq.heappop(queue)
                print(f"Kuyruk: {queue}")  # Kuyruk her adımda yazdırılacak
                for neighbor in graph[current_node]:
                    weight = graph[current_node][neighbor]['weight']
                    distance = current_distance + weight
                    if neighbor not in distances or distance < distances[neighbor]:
                        distances[neighbor] = distance
                        previous_nodes[neighbor] = current_node
                        heapq.heappush(queue, (distance, neighbor))

        # Kuyruğu adım adım gösterme
        print_queue_steps(G, A, B)

        # A yazarı ile işbirliği yaptığı yazarları listeleme
        def get_collaborators(graph, author):
            collaborators = list(graph.neighbors(author))
            print(f"{author} yazarı ile işbirliği yapan yazarlar: {collaborators}")
            return collaborators

        # A yazarıyla işbirliği yapan yazarları listeleme
        get_collaborators(G, A)


    
    """class GraphVisualization:
        def __init__(self, graph):
            self.graph = graph

        def yazarlar_arasi_en_kisa_yol(self):
        # Kullanıcıdan yazar ID'leri alınır
            yazar_a = simpledialog.askstring("Yazar A ID", "Lütfen A yazarının ID'sini giriniz:")
            yazar_b = simpledialog.askstring("Yazar B ID", "Lütfen B yazarının ID'sini giriniz:")

            if yazar_a not in self.graph.nodes or yazar_b not in self.graph.nodes:
                print(f"Hata: '{yazar_a}' veya '{yazar_b}' grafikte mevcut değil!")
                return

            try:
            # A ve B arasındaki en kısa yol ve mesafeyi hesapla
                yol = nx.dijkstra_path(self.graph, source=yazar_a, target=yazar_b)
                mesafe = nx.dijkstra_path_length(self.graph, source=yazar_a, target=yazar_b)
            except nx.NetworkXNoPath:
                print(f"Hata: '{yazar_a}' ile '{yazar_b}' arasında bir bağlantı yok!")
                return

        # Kuyruk içeriğini adım adım göstermek için PriorityQueue simülasyonu
            pq = PriorityQueue()
            pq.put((0, [yazar_a]))

            print("Kuyruk Adımları:")
            while not pq.empty():
                maliyet, yol_sozde = pq.get()
                print(f"Maliyet: {maliyet}, Yol: {yol_sozde}")

                if yol_sozde[-1] == yazar_b:
                    break

                for komsu, veri in self.graph[yol_sozde[-1]].items():
                    if komsu not in yol_sozde:
                        yeni_maliyet = maliyet + veri['weight']
                        yeni_yol = yol_sozde + [komsu]
                        pq.put((yeni_maliyet, yeni_yol))

        # Grafiksel görselleştirme için renkler belirle
            renkler = ['purple' if node == yazar_a or node == yazar_b else 'orchid' for node in self.graph.nodes]
            kenar_renkleri = ['green' if edge in zip(yol, yol[1:]) else 'gray' for edge in self.graph.edges]

        # Grafik ve tabloyu aynı anda göstermek
            fig, ax = plt.subplots(1, 2, figsize=(20, 8))

        # Grafik oluşturma
            pos = nx.spring_layout(self.graph)
            nx.draw(
                self.graph, pos, with_labels=True, node_color=renkler, edge_color=kenar_renkleri,
                node_size=500, font_size=10, font_color='black', ax=ax[0]
        )
            ax[0].set_title(f"'{yazar_a}' ile '{yazar_b}' Arasındaki En Kısa Yol")

        # Yol tablosu oluşturma
            veri = {'Yol Üzerindeki Yazarlar': yol, 'Toplam Mesafe': [mesafe] * len(yol)}
            tablo = pd.DataFrame(veri)

        # Tabloyu ekranda göstermek
            ax[1].axis('off')
            ax[1].table(cellText=tablo.values, colLabels=tablo.columns, loc='center')

            plt.tight_layout()
            plt.show()
"""

#2.ister
    def agirliga_gore_kuyruk_olustur(self):
        orcid = simpledialog.askstring("Yazar ID'si", "Lütfen Yazar ID'sini giriniz:")
        self.left_panel.update_text(f"Ağırlıklara Göre Kuyruk Oluşturuldu. Yazar ID: {orcid}")

        if orcid is None or orcid.strip() == "":
            messagebox.showerror("Hata!!", "Geçerli bir Yazar ID'si girilmedi!!")
            return

        orcid = orcid.strip()

    # Excel dosyasını yükleme
        try:
            workbook = openpyxl.load_workbook("PROLAB 3.xlsx")
            sheet = workbook.active
        except Exception as e:
            messagebox.showerror("Hata!!", f"Excel dosyası açılamadı!!: {e}")
            return

    # Tüm işbirliği yapılan yazarları toplamak için bir dictionary kullanıyoruz
        isbirligi_agirliklari_listesi = {}

    # ID arama işlemi ve işbirliklerini bulma
        for satir in sheet.iter_rows(min_row=2):  # Başlık satırını atlayarak işlem yap
            if str(satir[0].value) == orcid:  # A sütununda yazar ID'si aranır
                coauthors = satir[4].value  # E sütunundaki "coauthors" bilgisi alınır (index 4)
                if coauthors:
                    isbirliklerini_temizle = coauthors.replace("[", "").replace("]", "").replace("\"", "")
                    isbirligi_listesi = [name.strip() for name in isbirliklerini_temizle.split(",") if name.strip()]
                    for isbirlikci in isbirligi_listesi:
                        weight = isbirligi_agirliklari_listesi.get(isbirlikci, 0)
                        isbirligi_agirliklari_listesi[isbirlikci] = weight + 1  # Ağırlık sadece 1 eklenir

        if not isbirligi_agirliklari_listesi:
            messagebox.showerror("Hata!!", f"ID {orcid} Excel dosyasında bulunamadı!!!")
            return

    # Kuyruğa ağırlıklarına göre sıraya ekleme
        isbirlikcileri_siralama = sorted(isbirligi_agirliklari_listesi.items(), key=lambda item: item[1], reverse=True)

        for isbirlikci, weight in isbirlikcileri_siralama:
            self.left_panel.update_text(f"{isbirlikci}: {weight} ağırlık")

        messagebox.showinfo("Kuyruk Oluşturuldu", f"Ağırlıklara Göre Kuyruk oluşturuldu. Toplam {len(isbirlikcileri_siralama)} işbirlikçi yazar bulundu.")


#3. ister
    def bst_olustur(self):
        author_id = simpledialog.askstring("Yazar ID'si", "Lütfen Yazar ID'sini giriniz:")
        self.left_panel.update_text(f"Ağırlıklara Göre Kuyruk Oluşturuldu. Yazar ID: {author_id}")

        class BSTNode:
            def __init__(self, key, weight):
                self.key = key
                self.weight = weight
                self.left = None
                self.right = None

        class AuthorCollaboration:
            def __init__(self, excel_path, panel):
                self.excel_path = excel_path
                self.root = None
                self.panel = panel

            def read_excel(self):
        # Excel dosyasını okur ve DataFrame döner
                return pd.read_excel(self.excel_path, sheet_name='in')

            def isbirliklerini_hesapla(self, data, author_id):
        # İşbirliklerini hesaplar
                isbirlikleri = {}

        # Yazarın işbirliklerini bul
                author_data = data[data['orcid'] == author_id]

                for coauthors in author_data['coauthors']:
                    coauthors_listesi = eval(coauthors)  # Stringi listeye çeviriyoruz
                    for coauthor in coauthors_listesi:
                        if coauthor != author_id:
                            isbirlikleri[coauthor] = isbirlikleri.get(coauthor, 0) + 1

                return isbirlikleri

            def bst_ekle(self, root, key, weight):
        # İkili arama ağacına bir düğüm ekler
                if root is None:
                    return BSTNode(key, weight)
                if weight <= root.weight:
                    root.left = self.bst_ekle(root.left, key, weight)
                else:
                    root.right = self.bst_ekle(root.right, key, weight)
                return root

            def bst_olustur(self, author_id):
        # Excel dosyasını okuyarak ilgili yazarın işbirliklerini bulur
                data = self.read_excel()

        # İşbirliklerini hesapla
                isbirlikleri = self.isbirliklerini_hesapla(data, author_id)

                if not isbirlikleri:
                    self.panel.update_text("Belirtilen yazar ID'sine ait işbirliği bulunamadı.")
                    return

        # Öncelikli kuyruk (ağırlık sırasına göre sıralama için)
                pq = PriorityQueue()
                for coauthor, weight in isbirlikleri.items():
                    pq.put((-weight, coauthor))  # Ağırlık sırasını tersine çevirerek öncelik veririz

        # Öncelikli kuyruktan BST oluşturma
                while not pq.empty():
                    weight, coauthor = pq.get()
                    self.root = self.bst_ekle(self.root, coauthor, -weight)  # Ağırlık tekrar pozitif yapılır

            def inorder_traversal(self, node):
        # BST'yi sıralı şekilde gezmek için
                if node is not None:
                    self.inorder_traversal(node.left)
                    self.panel.update_text(f"{node.key}:{node.weight} ağırlık")
                    self.inorder_traversal(node.right)

            def bst_to_anytree(self, node, parent=None):
    # BST'yi AnyTree formatına çevirir
                if node is None:
                    return None

    # Anahtar ve ağırlığı (veya herhangi bir değeri) göstermek için şekillendirme
                current = Node(f"{node.key} ({node.weight})" if node.weight else f"{node.key}", parent=parent)
                if node.left:
                    self.bst_to_anytree(node.left, current)
                if node.right:
                    self.bst_to_anytree(node.right, current)
                return current

            def visualize_bst(self):
    # BST'yi görselleştirir
                anytree_root = self.bst_to_anytree(self.root)
                for pre, fill, node in RenderTree(anytree_root):
                    self.panel.update_text(f"{pre}{node.name}")
                #DotExporter(anytree_root).to_picture("bst_tree.png")
                #self.panel.update_text("Ağaç görselleştirildi ve 'bst_tree.png' dosyasına kaydedildi.")
        
        root = tk.Tk()
        root.title("BST Görselleştirme")

        sol_panel_instance = sol_panel(root)
        
        excel_path = "PROLAB 3.xlsx"
        
        author_collab = AuthorCollaboration(excel_path, sol_panel_instance)
        #author_collab = AuthorCollaboration(excel_path)
        
        author_collab.bst_olustur(author_id)

        author_collab.inorder_traversal(author_collab.root)
        author_collab.visualize_bst()
        root.mainloop()
     
#4.ister
    def kisa_yollari_hesapla(self):
        # Kullanıcıdan yazar adi alınır
        yazar_ad = simpledialog.askstring("Yazar Adi", "Lütfen Yazar Adini giriniz:")
        
        if yazar_ad not in self.graph.nodes:
            print(f"Hata: '{yazar_ad}' grafikte mevcut değil!")
            return

    # Girilen yazarın bağlantılı olduğu tüm yazarlardan oluşan alt graf
        alt_graf = self.graph.subgraph(nx.single_source_shortest_path(self.graph, source=yazar_ad).keys())

    # En kısa yollar ve mesafeler
        yollar = nx.single_source_dijkstra_path(alt_graf, source=yazar_ad)
        mesafeler = nx.single_source_dijkstra_path_length(alt_graf, source=yazar_ad)

    # Renkli görselleştirme
        renkler = []
        for node in alt_graf.nodes:
            if node == yazar_ad:
                renkler.append('purple')
            elif node in yollar:
                renkler.append('orchid')
            else:
                renkler.append('red')

    # Grafik ve tabloyu aynı ekranda göstermek
        fig, ax = plt.subplots(1, 2, figsize=(20, 8))

    # Görselleştirme
        pos = nx.spring_layout(alt_graf)
        nx.draw(
            alt_graf, pos, with_labels=True, node_color=renkler,
            node_size=500, font_size=10, font_color='black', ax=ax[0]
    )
        ax[0].set_title(f"'{yazar_ad}' Yazarının Bağlantılarıyla Olan En Kısa Yollar")

    # Tablo oluşturma
        veri = {'Hedef Yazar': list(mesafeler.keys()), 'Mesafe': list(mesafeler.values())}
        tablo = pd.DataFrame(veri)

    # Tabloyu ekranda göstermek
        ax[1].axis('off')
        ax[1].table(cellText=tablo.values, colLabels=tablo.columns, loc='center')

        plt.tight_layout()
        plt.show()
    
#5.ister
    def toplam_yazar_sayisi_hesapla(self):
     # Kullanıcıdan Yazar ID'si (ORCID) girişi alınır
        orcid = simpledialog.askstring("Yazar ID'si", "Lütfen Yazar ID'sini (ORCID) giriniz:")
    
        if orcid is None or orcid.strip() == "":
            messagebox.showerror("Hata", "Geçerli bir Yazar ID'si girilmedi.")
            return
    
        orcid = orcid.strip()
    
    # Excel dosyasını yükleme
        try:
            workbook = openpyxl.load_workbook("PROLAB 3.xlsx")
            sheet = workbook.active
        except Exception as e:
            messagebox.showerror("Hata", f"Excel dosyası açılamadı: {e}")
            return
    
    # Tüm işbirliği yapılan yazarları toplamak için bir set kullanıyoruz
        tum_isbirlikleri = set()

    # ID arama işlemi ve işbirliklerini bulma
        for satir in sheet.iter_rows(min_row=2):  # Başlık satırını atlayarak işlem yap
            if str(satir[0].value) == orcid:  # A sütununda yazar ID'si aranır
                isbirlikleri = satir[4].value  # E sütunundaki "coauthors" bilgisi alınır (index 4)
                if isbirlikleri:
                # Tırnak ve köşeli parantezleri kaldır, ardından isimleri ayır
                    isbirliklerini_temizle = isbirlikleri.replace("[", "").replace("]", "").replace("\"", "")
                    isbirligi_listesi = [name.strip() for name in isbirliklerini_temizle.split(",") if name.strip()]
                    tum_isbirlikleri.update(isbirligi_listesi)  # Set ile benzersiz yazarlar eklenir

        if not tum_isbirlikleri:
            messagebox.showerror("Hata!!", f"ID {orcid} Excel dosyasında bulunamadı!!!")
            return
    
    # Toplam işbirliği yapılan yazar sayısını hesapla
        toplam_isbirlikleri = len(tum_isbirlikleri)
    
    # Sonucu kullanıcıya göster
        isbirligi_listesi = sorted(tum_isbirlikleri)  # İsimleri alfabetik sıraya koy
        message = (
        f"ID {orcid} için işbirliği yapılan toplam yazar sayısı: {toplam_isbirlikleri}\n"
        f"İşbirliği yapılan yazarlar:\n{', '.join(isbirligi_listesi)}"
    )
        self.left_panel.update_text(message)
        messagebox.showinfo("Sonuç", message)


#6.ister
    def en_cok_isbirligi_yapan_yazar(self):
    # Excel dosyasını yükle
            wb = openpyxl.load_workbook("PROLAB 3.xlsx")
            sheet = wb.active

        # ORCID'ler ve yazar isimlerini oku
            orcid_ve_yazar = {}
            for satir in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
                orcid, author_name = satir[0], satir[3]
                if orcid and author_name:
                    orcid_ve_yazar[orcid] = author_name

        # Ağırlıklı işbirliği grafiği oluştur
            graph = defaultdict(int)
            for satir in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
                orcid1, orcid2, weight = satir[0], satir[1], satir[2]
                if orcid1 and orcid2 and weight:
                    graph[(orcid1, orcid2)] += weight

        # İşbirliği sayısını hesapla
            isbirligi_sayisi = defaultdict(int)
            for (orcid1, orcid2), weight in graph.items():
                isbirligi_sayisi[orcid1] += weight
                isbirligi_sayisi[orcid2] += weight

        # En çok işbirliği yapan yazarı bul
            en_cok_isbirligi_orcidi = max(isbirligi_sayisi, key=isbirligi_sayisi.get)
            max_isbirligi = isbirligi_sayisi[en_cok_isbirligi_orcidi]
            en_isbirlikci_yazar = orcid_ve_yazar.get(en_cok_isbirligi_orcidi, "Bilinmiyor")

        # Sonucu ekrana yazdır
            message = (
            f"En çok işbirliği yapan yazar: {en_isbirlikci_yazar}"
            f"\nToplam işbirliği sayısı: {max_isbirligi}"
        )
            print(message)
            messagebox.showinfo("En Çok İşbirliği Yapan Yazar", message)
            
            self.left_panel.update_text(message)

#7.ister
    def en_uzun_yolu_bul(self):
       orcid = simpledialog.askstring("Yazar Adi", "Lütfen Yazar Adi giriniz:")
       if orcid is None or orcid.strip() == "":
           messagebox.showerror("Hata!!", "Geçerli bir Yazar Adi girilmedi!!")
           return
    
       orcid = orcid.strip()
    
       if orcid not in graph.nodes:
           messagebox.showerror("Hata!!", f"ID {orcid} bulunamadı!!")
           return
    
    # BFS ile en uzun yolu bulma
       ziyaret_edilen_komsu = set()
       kuyruk = [(orcid, [orcid])]  # (Yazar ID, Yol Listesi)
       en_uzun_yol = (None, [])  # (Uzunluk, Yol)
    
       while kuyruk:
           mevcut_id, yol = kuyruk.pop(0)
        
           if mevcut_id not in ziyaret_edilen_komsu:
               ziyaret_edilen_komsu.add(mevcut_id)
            
               for komsu in graph.neighbors(mevcut_id):
                   if komsu not in ziyaret_edilen_komsu:
                        yeni_yol = yol + [komsu]
                        kuyruk.append((komsu, yeni_yol))
                        if len(yeni_yol) > len(en_uzun_yol[1]):
                          en_uzun_yol = (len(yeni_yol), yeni_yol)
    
       if en_uzun_yol[0] == 1:
           messagebox.showinfo("Sonuç", f"Yazar Adi {orcid} için en uzun yol: {en_uzun_yol[1]}")
       else:
           messagebox.showinfo("Sonuç", f"Yazar Adi {orcid} için en uzun yol: {en_uzun_yol[1]} (Uzunluk: {en_uzun_yol[0]})")

       self.left_panel.update_text(f"En uzun yol bulundu. Yazar Adi: {orcid}")


# Excel dosyasından verileri al ve grafiği çiz
file_name = "PROLAB 3.xlsx"  # Excel dosyanızın adını burada belirtin
graph, article_counts = excelden_graf_olusturma(file_name)

# Ana pencere oluşturma
root = tk.Tk()
root.title("Yazarlar ve İşbirliği Analizi")

# Sol ve sağ panelleri oluşturma
left_panel = sol_panel(root)
right_panel = sag_panel(root, left_panel,graph)


wb = load_workbook(filename="PROLAB 3.xlsx")
sheet = wb.active
graf_gorsellestirme(graph, article_counts, root)

# Belirli bir boyutta ayarla
root.geometry("1500x900")

# Pencereyi tam ekran başlatma
root.mainloop()